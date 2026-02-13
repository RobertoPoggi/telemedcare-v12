#!/usr/bin/env python3
"""
Aggiorna gli stati dei 12 lead basandosi sul foglio Excel (solo se non già valorizzati)
"""
import requests
import openpyxl

# Mappa Esito Excel → Stato DB
ESITO_TO_STATO = {
    'Interessata': 'Interessato',
    'Interessato': 'Interessato',
    'Non risponde': 'Da Ricontattare',
    'Da ricontattare': 'Da Ricontattare',
    'Non interessata': 'Non Interessato',
    'Inviata': 'Contattato',
    'In trattativa': 'In Trattativa',
    'Convertito': 'Convertito'
}

API_BASE = 'https://telemedcare-v12.pages.dev'

# Carica foglio Excel
wb = openpyxl.load_workbook('REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx')
ws = wb.active
header_row = 5

# Mappa nome → esito dal foglio
excel_esiti = {}

for row_idx in range(header_row + 1, ws.max_row + 1):
    nome = ws.cell(row_idx, 4).value  # Colonna D
    esito = ws.cell(row_idx, 7).value  # Colonna G
    
    if nome and esito:
        nome_norm = str(nome).strip().lower()
        esito_clean = str(esito).strip()
        excel_esiti[nome_norm] = esito_clean

print(f"📋 Caricati {len(excel_esiti)} esiti dal foglio Excel\n")

# I 12 lead importati
IMPORTED_LEADS_IDS = [
    'LEAD-MANUAL-1770946050424',  # Laura Bianchi
    'LEAD-MANUAL-1770946051660',  # Paola Scarpin
    'LEAD-MANUAL-1770946054097',  # Adriana Mulassano
    'LEAD-MANUAL-1770946055276',  # Maria Chiara Baldassini
    'LEAD-MANUAL-1770946057712',  # Andrea Mercuri
    'LEAD-MANUAL-1770946058870',  # Enzo Pedron
    'LEAD-MANUAL-1770946060015',  # Andrea Dindo
    'LEAD-MANUAL-1770946061218',  # Francesco Egiziano
    'LEAD-MANUAL-1770946062533',  # Mary De Sanctis
    'LEAD-MANUAL-1770946063740',  # Giovanna Giordano
    'LEAD-MANUAL-1770946064859',  # Marco Olivieri
    'LEAD-MANUAL-1770946067171'   # Alberto Avanzi
]

# Scarica i lead dal DB
print("📥 Caricamento lead dal database...")
response = requests.get(f'{API_BASE}/api/leads?limit=500', timeout=30)
all_leads = response.json().get('leads', [])

print(f"✅ Caricati {len(all_leads)} lead\n")

print(f"{'='*110}")
print(f"🔄 AGGIORNAMENTO STATI DA FOGLIO EXCEL")
print(f"{'='*110}\n")

updated = []
skipped_already_set = []
skipped_not_found = []

for lead_id in IMPORTED_LEADS_IDS:
    # Trova il lead nel DB
    lead = next((l for l in all_leads if l.get('id') == lead_id), None)
    
    if not lead:
        print(f"⚠️  Lead {lead_id[:30]} non trovato nel DB")
        skipped_not_found.append(lead_id)
        continue
    
    nome_completo = f"{lead.get('nomeRichiedente', '')} {lead.get('cognomeRichiedente', '')}".strip()
    stato_attuale = lead.get('stato')
    
    print(f"📋 {nome_completo:30} | Stato attuale: {stato_attuale or 'NULL'}")
    
    # Se stato già valorizzato, salta
    if stato_attuale and stato_attuale not in ['', 'null', 'None', 'Nuovo']:
        print(f"   ⏩ Stato già valorizzato, skip\n")
        skipped_already_set.append(nome_completo)
        continue
    
    # Cerca l'esito nel foglio Excel
    nome_norm = nome_completo.lower().strip()
    esito_excel = None
    
    # Cerca match esatto o parziale
    for excel_nome, esito in excel_esiti.items():
        cognome_lead = lead.get('cognomeRichiedente', '').lower()
        if cognome_lead and cognome_lead in excel_nome:
            esito_excel = esito
            break
    
    if esito_excel:
        nuovo_stato = ESITO_TO_STATO.get(esito_excel)
        
        if nuovo_stato:
            print(f"   📊 Esito Excel: {esito_excel} → Stato: {nuovo_stato}")
            print(f"   🔄 Aggiornamento...")
            
            try:
                response = requests.put(
                    f'{API_BASE}/api/leads/{lead_id}',
                    json={'stato': nuovo_stato},
                    timeout=10
                )
                
                if response.status_code in [200, 201]:
                    print(f"   ✅ Aggiornato\n")
                    updated.append({
                        'nome': nome_completo,
                        'esito': esito_excel,
                        'nuovo_stato': nuovo_stato
                    })
                else:
                    print(f"   ❌ Errore {response.status_code}\n")
            except Exception as e:
                print(f"   ❌ Errore: {str(e)[:50]}\n")
        else:
            print(f"   ⚠️  Esito '{esito_excel}' non mappato\n")
    else:
        print(f"   ⚠️  Esito non trovato nel foglio Excel\n")

# Report finale
print(f"{'='*110}")
print(f"📊 RISULTATI AGGIORNAMENTO STATI:")
print(f"{'='*110}")
print(f"✅ Aggiornati: {len(updated)}")
print(f"⏩ Già valorizzati (skip): {len(skipped_already_set)}")
print(f"⚠️  Non trovati in DB: {len(skipped_not_found)}")
print(f"{'='*110}\n")

if updated:
    print(f"✅ STATI AGGIORNATI:\n")
    for item in updated:
        print(f"   ✓ {item['nome']:30} | Esito: {item['esito']:20} → Stato: {item['nuovo_stato']}")

if skipped_already_set:
    print(f"\n⏩ STATI GIÀ VALORIZZATI (non modificati):\n")
    for nome in skipped_already_set:
        print(f"   - {nome}")

print(f"\n{'='*110}\n")
