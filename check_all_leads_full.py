#!/usr/bin/env python3
"""
Script per verificare tutti i lead nel DB contro il foglio Excel di Ottavia (senza limite)
"""
import json
import sqlite3
from pathlib import Path
import openpyxl
from datetime import datetime

# Percorsi
DB_PATH = '/home/user/webapp/.wrangler/state/v3/d1/miniflare-D1DatabaseObject/ddb6129a68cd0fdb2e8cc21a5afd3e8bfedfb5db6dbbeab62b9f11dcc7a9f39d.sqlite'
EXCEL_FILE = '/home/user/webapp/REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx'

def normalize_text(text):
    """Normalizza testo per comparazione"""
    if not text:
        return ''
    return str(text).strip().lower()

def normalize_phone(phone):
    """Normalizza numero di telefono"""
    if not phone:
        return ''
    s = str(phone).strip()
    # Rimuove spazi, trattini, parentesi, +39, prefisso internazionale
    s = s.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    s = s.replace('+39', '').replace('0039', '')
    # Prende solo gli ultimi 9-10 cifre
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) >= 9:
        return digits[-10:]  # Ultime 10 cifre
    return digits

def load_db_leads():
    """Carica TUTTI i lead dal database"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # RIMUOVO IL LIMITE - prendi TUTTI i lead
    cursor.execute("""
        SELECT id, richiedente, email, telefono
        FROM leads
        ORDER BY created_at DESC
    """)
    
    leads = []
    for row in cursor.fetchall():
        leads.append({
            'id': row['id'],
            'name': row['richiedente'],
            'email': row['email'],
            'telefono': row['telefono']
        })
    
    conn.close()
    print(f"✅ Caricati {len(leads)} lead dal database")
    
    # Statistiche
    unique_names = len(set(normalize_text(l['name']) for l in leads if l['name']))
    unique_emails = len(set(normalize_text(l['email']) for l in leads if l['email']))
    unique_phones = len(set(normalize_phone(l['telefono']) for l in leads if l['telefono']))
    
    print(f"   - {unique_names} nomi unici")
    print(f"   - {unique_emails} email uniche")
    print(f"   - {unique_phones} telefoni unici")
    
    return leads

def load_excel_leads():
    """Carica i lead dal foglio Excel di Ottavia"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Trova la riga header
    header_row = None
    for i in range(1, 10):
        cell_val = ws.cell(i, 1).value
        if cell_val and 'DATA' in str(cell_val).upper():
            header_row = i
            break
    
    if not header_row:
        raise ValueError("Header non trovato nel file Excel")
    
    print(f"✅ Header trovato alla riga {header_row}")
    
    # Leggi tutti i lead
    excel_leads = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Colonne: A=DATA, D=NOME, E=CONTATTO
        data = ws.cell(row_idx, 1).value
        nome = ws.cell(row_idx, 4).value
        contatto = ws.cell(row_idx, 5).value
        tipo_attivita = ws.cell(row_idx, 6).value
        esito = ws.cell(row_idx, 7).value
        
        # Salta righe vuote o separatori
        if not nome or not contatto:
            continue
        if 'TOTALE' in str(nome).upper():
            continue
            
        excel_leads.append({
            'row': row_idx,
            'data': data,
            'nome': nome,
            'contatto': contatto,
            'tipo': tipo_attivita,
            'esito': esito
        })
    
    print(f"✅ Caricati {len(excel_leads)} lead dall'Excel")
    return excel_leads

def find_matching_lead(excel_lead, db_leads):
    """Cerca un lead nel DB che corrisponda al lead Excel"""
    excel_nome = normalize_text(excel_lead['nome'])
    excel_contatto = str(excel_lead['contatto']).strip()
    
    # Determina se il contatto è email o telefono
    is_email = '@' in excel_contatto
    
    if is_email:
        excel_email = normalize_text(excel_contatto)
        # Match per email
        for db_lead in db_leads:
            if normalize_text(db_lead['email']) == excel_email:
                return db_lead
    else:
        excel_phone = normalize_phone(excel_contatto)
        # Match per telefono
        for db_lead in db_leads:
            if normalize_phone(db_lead['telefono']) == excel_phone:
                return db_lead
    
    # Match per nome (fallback)
    for db_lead in db_leads:
        if normalize_text(db_lead['name']) == excel_nome:
            return db_lead
    
    return None

def main():
    print("\n" + "="*80)
    print("🔍 VERIFICA COMPLETA LEAD MANCANTI (SENZA LIMITI)")
    print("="*80 + "\n")
    
    # Carica dati
    db_leads = load_db_leads()
    excel_leads = load_excel_leads()
    
    # Trova lead mancanti
    missing_leads = []
    matched_count = 0
    
    for excel_lead in excel_leads:
        match = find_matching_lead(excel_lead, db_leads)
        if match:
            matched_count += 1
        else:
            missing_leads.append(excel_lead)
    
    # Risultati
    print("\n" + "="*80)
    print(f"📊 RISULTATI:")
    print(f"   - Lead nel DB: {len(db_leads)}")
    print(f"   - Lead nell'Excel: {len(excel_leads)}")
    print(f"   - Lead trovati: {matched_count}")
    print(f"   - Lead MANCANTI: {len(missing_leads)}")
    print(f"   - % Match: {(matched_count/len(excel_leads)*100):.1f}%")
    print("="*80 + "\n")
    
    if missing_leads:
        print(f"\n❌ LEAD MANCANTI NEL DATABASE ({len(missing_leads)}):\n")
        for i, lead in enumerate(missing_leads, 1):
            print(f"{i:3}. Riga {lead['row']:3} | {lead['nome']:30} | {lead['contatto']:25} | {lead['tipo']:12} | {lead['esito']:20} | {lead['data']}")
        
        # Salva in JSON
        output_file = '/home/user/webapp/leads_mancanti_full.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(missing_leads, f, indent=2, ensure_ascii=False, default=str)
        print(f"\n💾 Lista completa salvata in: {output_file}")
    else:
        print("\n✅ Tutti i lead dell'Excel sono presenti nel database!")
    
    print("\n" + "="*80 + "\n")

if __name__ == '__main__':
    main()
