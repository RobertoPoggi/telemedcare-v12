#!/usr/bin/env python3
"""
Script per importare le interazioni di Ottavia dal file Excel
NON tocca Andrea D'Avella (già inserito manualmente)
"""

import openpyxl
import json
from datetime import datetime
import sys

def convert_excel_date(excel_date):
    """Converte data Excel in formato ISO"""
    if not excel_date or excel_date == '-':
        return None
    
    if isinstance(excel_date, datetime):
        return excel_date.strftime('%Y-%m-%d')
    
    # Prova formato DD/MM/YYYY
    try:
        if isinstance(excel_date, str) and '/' in excel_date:
            parts = excel_date.split('/')
            if len(parts) == 3:
                day, month, year = parts
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    except:
        pass
    
    return None

def main():
    excel_path = '/home/user/uploaded_files/REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx'
    
    # Carica workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Tracker Giornaliero']
    
    # Trova header (riga 5)
    header_row = 5
    
    # Colonne secondo le specifiche:
    # A=DATA, B=GIORNO, C=CANALE, D=NOME/AZIENDA, E=CONTATTO
    # F=TIPO ATTIVITÀ, G=ESITO, H=PROSSIMO STEP, I=DATA FOLLOW-UP, J=NOTE
    
    interactions = []
    skipped = 0
    davella_skipped = 0
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Leggi valori
        data_val = ws.cell(row_idx, 1).value  # A - DATA
        nome_val = ws.cell(row_idx, 4).value  # D - NOME/AZIENDA
        contatto_val = ws.cell(row_idx, 5).value  # E - CONTATTO
        tipo_val = ws.cell(row_idx, 6).value  # F - TIPO ATTIVITÀ
        esito_val = ws.cell(row_idx, 7).value  # G - ESITO
        step_val = ws.cell(row_idx, 8).value  # H - PROSSIMO STEP
        followup_val = ws.cell(row_idx, 9).value  # I - DATA FOLLOW-UP
        note_val = ws.cell(row_idx, 10).value  # J - NOTE
        
        # Skip righe vuote o separatori
        if not nome_val or nome_val == '-' or str(nome_val).startswith('sett'):
            skipped += 1
            continue
        
        # Skip Andrea D'Avella (già inserito manualmente)
        nome_str = str(nome_val).strip()
        if "D'Avella" in nome_str or "D'AVELLA" in nome_str.upper() or "DAVELLA" in nome_str.upper():
            print(f"⏭️  Skip Andrea D'Avella (già inserito manualmente)")
            davella_skipped += 1
            continue
        
        # Converti data
        data_iso = convert_excel_date(data_val)
        
        # Costruisci nota completa
        nota_parts = []
        if esito_val and str(esito_val).strip() != '-':
            nota_parts.append(f"Esito: {str(esito_val).strip()}")
        if note_val and str(note_val).strip() != '-':
            nota_parts.append(f"Note: {str(note_val).strip()}")
        
        nota_completa = '. '.join(nota_parts) if nota_parts else 'Contatto effettuato'
        
        # Azione = PROSSIMO STEP
        azione = str(step_val).strip() if step_val and str(step_val).strip() != '-' else None
        
        interaction = {
            'data': data_iso or '2026-01-12',  # Default se manca
            'nome_azienda': nome_str,
            'contatto': str(contatto_val).strip() if contatto_val else None,
            'tipo_attivita': str(tipo_val).strip() if tipo_val else 'Telefonata',
            'esito': str(esito_val).strip() if esito_val and str(esito_val).strip() != '-' else None,
            'prossimo_step': azione,
            'note': nota_completa
        }
        
        interactions.append(interaction)
    
    print(f"✅ Estratte {len(interactions)} interazioni")
    print(f"⏭️  Saltate {davella_skipped} righe di D'Avella")
    print(f"📝 Saltate {skipped} righe vuote/separatori")
    
    # Salva JSON
    output_data = {
        'operatore': 'Ottavia Belfa',
        'interactions': interactions
    }
    
    output_path = '/home/user/webapp/ottavia_interactions_import.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ File JSON creato: {output_path}")
    print(f"\n📤 Per importare, esegui:")
    print(f"curl -X POST https://telemedcare-v12.pages.dev/api/admin/import-interactions-json \\")
    print(f"  -H 'Content-Type: application/json' \\")
    print(f"  -d @{output_path}")
    
    # Mostra prime 5 interazioni
    print("\n📋 Prime 5 interazioni:")
    for i, int_data in enumerate(interactions[:5], 1):
        print(f"{i}. {int_data['nome_azienda']} - {int_data['tipo_attivita']} - {int_data['esito']}")

if __name__ == '__main__':
    main()
