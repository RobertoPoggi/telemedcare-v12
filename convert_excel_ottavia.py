#!/usr/bin/env python3
"""
Script per convertire il file Excel di Ottavia in formato JSON per l'import
Legge il foglio "Tracker Giornaliero" e estrae:
- DATA, GIORNO, CANALE, NOME/AZIENDA, CONTATTO, TIPO ATTIVITÀ, ESITO, PROSSIMO STEP, DATA FOLLOW-UP, NOTE

Output: interactions.json per l'import via API
"""

import openpyxl
import json
from datetime import datetime
import sys

def convert_excel_to_json(excel_path: str, output_path: str = 'interactions.json'):
    """
    Converte il file Excel di Ottavia in JSON per l'import
    """
    try:
        # Carica il workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Seleziona il foglio "Tracker Giornaliero"
        if "Tracker Giornaliero" not in wb.sheetnames:
            print(f"❌ Errore: Foglio 'Tracker Giornaliero' non trovato")
            print(f"Fogli disponibili: {wb.sheetnames}")
            return False
        
        ws = wb["Tracker Giornaliero"]
        
        # Trova la riga dell'header (di solito riga 5)
        header_row = None
        for row_idx in range(1, 10):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and str(cell_value).strip().upper() == 'DATA':
                header_row = row_idx
                break
        
        if not header_row:
            print("❌ Errore: Header non trovato nelle prime 10 righe")
            return False
        
        print(f"✅ Header trovato alla riga {header_row}")
        
        # Leggi l'header
        headers = []
        for col_idx in range(1, 15):  # Leggi fino a colonna 15
            cell_value = ws.cell(header_row, col_idx).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(None)
        
        print(f"Colonne trovate: {[h for h in headers if h]}")
        
        # Mapping colonne
        col_map = {}
        for idx, h in enumerate(headers):
            if h:
                h_upper = h.upper()
                if h_upper == 'DATA':
                    col_map['DATA'] = idx + 1
                elif 'NOME' in h_upper or 'AZIENDA' in h_upper:
                    col_map['NOME_AZIENDA'] = idx + 1
                elif 'CONTATTO' in h_upper:
                    col_map['CONTATTO'] = idx + 1
                elif 'TIPO' in h_upper and 'ATTIVITÀ' in h_upper:
                    col_map['TIPO_ATTIVITA'] = idx + 1
                elif h_upper == 'ESITO':
                    col_map['ESITO'] = idx + 1
                elif 'PROSSIMO' in h_upper or 'STEP' in h_upper:
                    col_map['PROSSIMO_STEP'] = idx + 1
                elif h_upper == 'NOTE':
                    col_map['NOTE'] = idx + 1
        
        print(f"Mappatura colonne: {col_map}")
        
        # Estrai le righe di dati
        interactions = []
        skipped = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Leggi i valori
            data_val = ws.cell(row_idx, col_map.get('DATA', 1)).value
            nome_val = ws.cell(row_idx, col_map.get('NOME_AZIENDA', 4)).value
            contatto_val = ws.cell(row_idx, col_map.get('CONTATTO', 5)).value
            tipo_val = ws.cell(row_idx, col_map.get('TIPO_ATTIVITA', 6)).value
            esito_val = ws.cell(row_idx, col_map.get('ESITO', 7)).value
            step_val = ws.cell(row_idx, col_map.get('PROSSIMO_STEP', 8)).value
            note_val = ws.cell(row_idx, col_map.get('NOTE', 10)).value
            
            # Skip righe vuote
            if not nome_val and not contatto_val:
                skipped += 1
                continue
            
            # Converti data
            data_str = None
            if data_val:
                if isinstance(data_val, datetime):
                    data_str = data_val.strftime('%Y-%m-%d')
                else:
                    try:
                        # Prova formato DD/MM/YYYY
                        parts = str(data_val).split('/')
                        if len(parts) == 3:
                            data_str = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                    except:
                        pass
            
            interaction = {
                'data': data_str,
                'nome_azienda': str(nome_val).strip() if nome_val else None,
                'contatto': str(contatto_val).strip() if contatto_val else None,
                'tipo_attivita': str(tipo_val).strip() if tipo_val else 'Telefonata',
                'esito': str(esito_val).strip() if esito_val else None,
                'prossimo_step': str(step_val).strip() if step_val else None,
                'note': str(note_val).strip() if note_val else None
            }
            
            interactions.append(interaction)
        
        print(f"✅ Estratte {len(interactions)} interazioni (saltate {skipped} righe vuote)")
        
        # Salva in JSON
        output_data = {
            'operatore': 'Ottavia Belfa',
            'interactions': interactions
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        print(f"✅ File JSON salvato: {output_path}")
        print(f"\n📤 Per importare le interazioni, esegui:")
        print(f"curl -X POST https://telemedcare-v12.pages.dev/api/admin/import-interactions-json \\")
        print(f"  -H 'Content-Type: application/json' \\")
        print(f"  -d @{output_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Errore: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python convert_excel_ottavia.py <path_to_excel>")
        print("Esempio: python convert_excel_ottavia.py REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'interactions.json'
    
    success = convert_excel_to_json(excel_path, output_path)
    sys.exit(0 if success else 1)
