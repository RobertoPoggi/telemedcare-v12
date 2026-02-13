#!/usr/bin/env python3
"""
Script per verificare i lead mancanti usando l'API Cloudflare
"""
import json
import requests
import openpyxl
from datetime import datetime

# Endpoint API
API_URL = 'https://telemedcare-v12.pages.dev/api/leads?limit=500'  # Aumento il limite
EXCEL_FILE = '/home/user/webapp/REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx'

print(f"📁 Verifico esistenza file Excel: {EXCEL_FILE}")
import os
if not os.path.exists(EXCEL_FILE):
    print(f"❌ File non trovato: {EXCEL_FILE}")
    print("🔍 Tento percorso alternativo...")
    EXCEL_FILE = '/home/user/uploaded_files/REPORT_SETTIMANALE_Ottavia_TEMPLATE.xlsx'
    if os.path.exists(EXCEL_FILE):
        print(f"✅ Trovato in: {EXCEL_FILE}")
    else:
        raise FileNotFoundError(f"File Excel non trovato in nessuna posizione")

def normalize_text(text):
    """Normalizza testo per comparazione"""
    if not text:
        return ''
    return str(text).strip().lower()

def normalize_phone(phone):
    """Normalizza numero di telefono"""
    if not phone:
        return ''
    s = str(phone).strip()
    # Rimuove spazi, trattini, parentesi, +39, prefisso internazionale
    s = s.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    s = s.replace('+39', '').replace('0039', '')
    # Prende solo gli ultimi 9-10 cifre
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) >= 9:
        return digits[-10:]  # Ultime 10 cifre
    return digits

def load_api_leads():
    """Carica tutti i lead dall'API"""
    print("📥 Scaricando lead dall'API...")
    
    try:
        response = requests.get(API_URL, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # L'API ritorna { "leads": [...], "counts": {...} }
        leads = data.get('leads', [])
        
        print(f"✅ Caricati {len(leads)} lead dall'API")
        
        # Statistiche
        unique_names = len(set(normalize_text(l.get('richiedente', '')) for l in leads if l.get('richiedente')))
        unique_emails = len(set(normalize_text(l.get('email', '')) for l in leads if l.get('email')))
        unique_phones = len(set(normalize_phone(l.get('telefono', '')) for l in leads if l.get('telefono')))
        
        print(f"   - {unique_names} nomi unici")
        print(f"   - {unique_emails} email uniche")
        print(f"   - {unique_phones} telefoni unici")
        
        return leads
        
    except Exception as e:
        print(f"❌ Errore durante il download dei lead: {e}")
        return []

def load_excel_leads():
    """Carica i lead dal foglio Excel di Ottavia"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Trova la riga header
    header_row = None
    for i in range(1, 10):
        cell_val = ws.cell(i, 1).value
        if cell_val and 'DATA' in str(cell_val).upper():
            header_row = i
            break
    
    if not header_row:
        raise ValueError("Header non trovato nel file Excel")
    
    print(f"✅ Header trovato alla riga {header_row}")
    
    # Leggi tutti i lead
    excel_leads = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Colonne: A=DATA, D=NOME, E=CONTATTO
        data = ws.cell(row_idx, 1).value
        nome = ws.cell(row_idx, 4).value
        contatto = ws.cell(row_idx, 5).value
        tipo_attivita = ws.cell(row_idx, 6).value
        esito = ws.cell(row_idx, 7).value
        prossimo_step = ws.cell(row_idx, 8).value
        data_followup = ws.cell(row_idx, 9).value
        note = ws.cell(row_idx, 10).value
        
        # Salta righe vuote o separatori
        if not nome or not contatto:
            continue
        if 'TOTALE' in str(nome).upper():
            continue
            
        excel_leads.append({
            'row': row_idx,
            'data': data,
            'nome': nome,
            'contatto': contatto,
            'tipo': tipo_attivita,
            'esito': esito,
            'prossimo_step': prossimo_step,
            'data_followup': data_followup,
            'note': note
        })
    
    print(f"✅ Caricati {len(excel_leads)} lead dall'Excel")
    return excel_leads

def find_matching_lead(excel_lead, api_leads):
    """Cerca un lead nell'API che corrisponda al lead Excel"""
    excel_nome = normalize_text(excel_lead['nome'])
    excel_contatto = str(excel_lead['contatto']).strip()
    
    # Determina se il contatto è email o telefono
    is_email = '@' in excel_contatto
    
    if is_email:
        excel_email = normalize_text(excel_contatto)
        # Match per email
        for api_lead in api_leads:
            if normalize_text(api_lead.get('email', '')) == excel_email:
                return api_lead
    else:
        excel_phone = normalize_phone(excel_contatto)
        # Match per telefono
        for api_lead in api_leads:
            if normalize_phone(api_lead.get('telefono', '')) == excel_phone:
                return api_lead
    
    # Match per nome (fallback)
    for api_lead in api_leads:
        if normalize_text(api_lead.get('richiedente', '')) == excel_nome:
            return api_lead
    
    return None

def main():
    print("\n" + "="*80)
    print("🔍 VERIFICA COMPLETA LEAD MANCANTI (API CLOUDFLARE)")
    print("="*80 + "\n")
    
    # Carica dati
    api_leads = load_api_leads()
    if not api_leads:
        print("❌ Impossibile caricare i lead dall'API")
        return
    
    excel_leads = load_excel_leads()
    
    # Trova lead mancanti
    missing_leads = []
    matched_count = 0
    
    for excel_lead in excel_leads:
        match = find_matching_lead(excel_lead, api_leads)
        if match:
            matched_count += 1
        else:
            missing_leads.append(excel_lead)
    
    # Risultati
    print("\n" + "="*80)
    print(f"📊 RISULTATI:")
    print(f"   - Lead nel DB (API): {len(api_leads)}")
    print(f"   - Lead nell'Excel: {len(excel_leads)}")
    print(f"   - Lead trovati: {matched_count}")
    print(f"   - Lead MANCANTI: {len(missing_leads)}")
    print(f"   - % Match: {(matched_count/len(excel_leads)*100):.1f}%")
    print("="*80 + "\n")
    
    if missing_leads:
        print(f"\n❌ LEAD MANCANTI NEL DATABASE ({len(missing_leads)}):\n")
        for i, lead in enumerate(missing_leads, 1):
            data_str = lead['data'].strftime('%d/%m/%Y') if isinstance(lead['data'], datetime) else str(lead['data'])
            print(f"{i:3}. Riga {lead['row']:3} | {lead['nome']:30} | {lead['contatto']:25} | {lead['tipo']:12} | {lead['esito']:20} | {data_str}")
        
        # Salva in JSON
        output_file = '/home/user/webapp/leads_mancanti_full.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(missing_leads, f, indent=2, ensure_ascii=False, default=str)
        print(f"\n💾 Lista completa salvata in: {output_file}")
    else:
        print("\n✅ Tutti i lead dell'Excel sono presenti nel database!")
    
    print("\n" + "="*80 + "\n")

if __name__ == '__main__':
    main()
